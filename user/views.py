from datetime import timedelta
from django.utils.timezone import localtime

from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType

from django.contrib.auth.views import (
    LoginView,
    LogoutView,
    PasswordChangeView,
    PasswordResetView,
    PasswordResetDoneView,
    PasswordResetConfirmView,
    PasswordResetCompleteView,
)

from django.views.generic import View, ListView, DetailView, TemplateView
from django.views.generic.edit import UpdateView, DeleteView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin

from django.contrib.auth.hashers import make_password
from django.utils.http import urlsafe_base64_decode
from django.utils.encoding import force_str
from django.contrib.auth.tokens import default_token_generator


from django.conf import settings
import csv
from django.http import HttpRequest, HttpResponse
from django.contrib import messages

from openpyxl import Workbook

from .forms import (
    LoginForm,
    UserAddForm,
    UserRegisterForm,
    RequestEmailVerificationForm,
    UserEditForm,
    CustomPasswordChangeForm,
    ForgotPasswordForm,
    ResetPasswordForm,
    UserGroupAddForm,
)

from django.db.models import Q

from .models import CustomUser
from common.mixins import MyPermissionRequiredMixin
from . import utils


class UserRegisterView(CreateView):
    """Registers a new user account"""

    model = CustomUser
    form_class = UserRegisterForm
    template_name = "user/register.html"
    success_url = reverse_lazy("user_register_success")

    def dispatch(self, request: HttpRequest, *args, **kwargs):
        if not settings.PUBLIC_REGISTRATION:
            return render(
                request,
                "forbidden.html",
                {"message": "Public registration has been disabled."},
                status=403,
            )
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = form.save()

        if settings.EMAIL_VERIFICATION_REQUIRED:
            utils.send_verification_email(self, user)
            return redirect("email_verification_sent")

        # Set a one-time flag in session
        self.request.session["email_verified_success"] = True
        return super().form_valid(form)


class UserRegisterSuccessView(TemplateView):
    """Registration success view"""

    template_name = "user/register_success.html"

    def dispatch(self, request, *args, **kwargs):
        # Check and remove one-time session flag
        if not request.session.pop("email_verified_success", False):
            return render(
                request,
                "forbidden.html",
                {"message": "Direct access to this page is not allowed."},
                status=403,
            )
        return super().dispatch(request, *args, **kwargs)


class RequestEmailVerificationView(View):
    """Request email verification view"""

    form_class = RequestEmailVerificationForm
    template_name = "user/request_email_verification.html"
    title = "Email not Verified"
    error_message = "Your account has been created but the email verification is pending. Check your email and click on the link to verify your email. If you did not received the verification email, enter your email below and request email verification."

    def get(self, request):
        form = self.form_class()
        return render(
            request,
            self.template_name,
            {"form": form, "title": self.title, "error_message": self.error_message},
        )

    def post(self, request):
        form = self.form_class(request.POST)

        if form.is_valid():
            email = form.cleaned_data.get("email")
            user = CustomUser.objects.filter(email=email).first()
            utils.send_verification_email(self, user)
            return redirect("email_verification_sent")

        return render(
            request,
            self.template_name,
            {"form": form, "title": self.title, "error_message": self.error_message},
        )


class EmailVerificationTokenFailView(RequestEmailVerificationView):
    """Invalid email verification token view"""

    title = "Invalid Token"
    error_message = "The email verification link is invalid or has expired. Enter your email below and request email verification again."


class EmailVerificationSentView(TemplateView):
    """Email verification sent view"""

    template_name = "user/email_verification_sent.html"

    def dispatch(self, request, *args, **kwargs):
        """
        Only allow access if the user was referred from Register, RequestEmailVerificationView or EmailVerificationTokenFailView.
        Otherwise, show an error page.
        """
        allowed_referers = [
            request.build_absolute_uri(reverse("register")),
            request.build_absolute_uri(reverse("request_email_verification")),
            request.build_absolute_uri(reverse("email_verification_token_fail")),
        ]

        referer = request.META.get("HTTP_REFERER")

        if not referer or not any(referer.startswith(url) for url in allowed_referers):
            return render(
                request,
                "forbidden.html",
                {"message": "Direct access to this page is not allowed."},
                status=403,
            )

        return super().dispatch(request, *args, **kwargs)


class EmailVerificationView(View):
    """Email verification view"""

    def get(self, request, uidb64, token):
        try:
            uid = urlsafe_base64_decode(uidb64).decode()
            user = get_object_or_404(CustomUser, pk=uid)
        except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist):
            user = None

        if user and default_token_generator.check_token(user, token):
            user.email_verified = True
            user.save()
            # Set a one-time flag in session
            request.session["email_verified_success"] = True
            return redirect("user_register_success")
        else:
            return redirect("email_verification_token_fail")


class UserLoginView(LoginView):
    """User login view"""

    template_name = "user/login.html"
    authentication_form = LoginForm

    def dispatch(self, request, *args, **kwargs):
        """Redirect logged-in users from the login page to the dashboard."""
        if request.user.is_authenticated:
            return redirect(reverse_lazy("dashboard"))  # Redirect to dashboard
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["PUBLIC_REGISTRATION"] = settings.PUBLIC_REGISTRATION
        return context

    def get_success_url(self):
        next_url = self.request.GET.get("next")

        if next_url:
            return next_url
        else:
            return super().get_success_url()

    def form_valid(self, form):

        if settings.EMAIL_VERIFICATION_REQUIRED:
            if not form.get_user().email_verified:  # type: ignore
                return redirect("request_email_verification")

        remember_me = form.cleaned_data.get("remember_me")

        if remember_me:
            self.request.session.set_expiry(timedelta(days=5).seconds)
        else:
            self.request.session.set_expiry(timedelta(hours=1).seconds)

        return super().form_valid(form)


class UserLogoutView(LogoutView):
    """User logout view"""

    def get_success_url(self):
        next_url = self.request.GET.get("next")

        if next_url:
            return next_url
        else:
            return super().get_success_url()  # type: ignore


class UserForgotPasswordView(PasswordResetView):
    """Forgot password view"""

    form_class = ForgotPasswordForm
    template_name = "user/forgot_password.html"
    email_template_name = "user/forgot_password_email.txt"
    html_email_template_name = "user/forgot_password_email.html"
    extra_email_context = {"PROJECT_NAME_SHORT": settings.PROJECT_NAME_SHORT}
    subject_template_name = "user/forgot_password_subject.txt"
    success_url = reverse_lazy("user_forgot_password_success")

    def form_valid(self, form):
        email = form.cleaned_data.get("email")
        if not CustomUser.objects.filter(email=email).exists():
            form.add_error("email", "Account with this email address does not exist.")
            return self.form_invalid(form)

        return super().form_valid(form)


class UserForgotPasswordSuccessView(PasswordResetDoneView):
    """Display success message after forgot password view"""

    template_name = "user/forgot_password_success.html"

    def dispatch(self, request, *args, **kwargs):
        """
        Only allow access if the user was referred from UserForgotPasswordView.
        Otherwise, show an error page.
        """
        allowed_referers = [
            request.build_absolute_uri(reverse("forgot_password")),
        ]

        referer = request.META.get("HTTP_REFERER")

        if not referer or not any(referer.startswith(url) for url in allowed_referers):
            return render(
                request,
                "forbidden.html",
                {"message": "Direct access to this page is not allowed."},
                status=403,
            )

        return super().dispatch(request, *args, **kwargs)


class UserResetPasswordView(PasswordResetConfirmView):
    """Allow user to reset the password"""

    template_name = "user/reset_password.html"
    form_class = ResetPasswordForm
    success_url = reverse_lazy("user_reset_password_success")
    token_fail_template_name = "user/reset_password_token_fail.html"

    def dispatch(self, request, uidb64=None, token=None, *args, **kwargs):

        if not uidb64 or not token:
            return render(request, self.token_fail_template_name)

        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = get_object_or_404(CustomUser, pk=uid)
        except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist):
            user = None

        if user is not None and default_token_generator.check_token(user, token):
            # Token is valid; proceed with normal flow
            self.user = user
            self.validlink = True
            return super().dispatch(
                request, uidb64=uidb64, token=token, *args, **kwargs
            )
        else:
            # Token is invalid; show custom template
            self.validlink = False
            self.user = None
            context = self.get_context_data()
            context["form"] = None
            context["validlink"] = False
            return render(request, self.token_fail_template_name, context)

    def form_valid(self, form):
        # Set one time session flag
        self.request.session["reset_password_success"] = True
        return super().form_valid(form)


class UserResetPasswordSuccessView(PasswordResetCompleteView):
    """Display success message after completion of reset password"""

    template_name = "user/reset_password_success.html"

    def dispatch(self, request, *args, **kwargs):
        # Check and remove one-time session flag
        if not request.session.pop("reset_password_success", False):
            return render(
                request,
                "forbidden.html",
                {"message": "Direct access to this page is not allowed."},
                status=403,
            )
        return super().dispatch(request, *args, **kwargs)


class UserListView(
    LoginRequiredMixin,
    MyPermissionRequiredMixin,
    ListView,
):
    """Displays list of users based on filters applied"""

    model = CustomUser
    template_name = "user/list.html"
    context_object_name = "users"
    permission_required = "user.can_view_others_profile"
    permission_denied_message = "You do not have permission to view this page"
    paginate_by = 10

    def get_paginate_by(self, queryset):
        page_size = self.request.GET.get("page_size")
        if page_size and page_size.isdigit():
            return int(page_size)
        return self.paginate_by

    def get_queryset(self):
        status = self.request.GET.get("status", "")
        role = self.request.GET.get("role", "")
        query = self.request.GET.get("q", "")
        queryset = (
            CustomUser.objects.select_related()
            .prefetch_related("groups", "user_permissions")
            .order_by("full_name")
        )
        queryset = utils.get_filtered_users(
            queryset, status=status, role=role, query=query
        )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["current_page_size"] = self.get_paginate_by(self.get_queryset())
        context["page_size_options"] = [5, 10, 20, 50, 100]
        context["status_options"] = {
            "< Status >": "",
            "Active": "active",
            "Disabled": "disabled",
        }
        context["role_options"] = {
            "< Role >": "",
            "Admin": "admin",
            "User": "user",
        }
        return context


class UserListExportAsXlsxView(LoginRequiredMixin, MyPermissionRequiredMixin, View):
    """Exports user list as .xlsx file"""

    permission_required = "user.can_view_others_profile"

    def get(self, request, *args, **kwargs):
        status = self.request.GET.get("status", "")
        role = self.request.GET.get("role", "")
        query = self.request.GET.get("q", "")
        users = CustomUser.objects.order_by("full_name")
        users = utils.get_filtered_users(users, status=status, role=role, query=query)

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Sheet1")
        ws.append(["Name", "Status", "Role", "Email", "Phone", "Gender", "Date Joined"])

        for user in users:
            ws.append(
                [
                    user.full_name,
                    user.is_active and "Active" or "Disabled",
                    user.is_superuser and "Admin" or "User",
                    user.email,
                    user.phone_number,
                    user.gender,
                    localtime(user.date_joined).strftime("%d-%m-%Y %I:%M:%S %p %Z"),
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="User List.xlsx"'

        wb.save(response)
        return response


class UserListExportAsCsvView(LoginRequiredMixin, MyPermissionRequiredMixin, View):
    """Exports user list as .csv file"""

    permission_required = "user.can_view_others_profile"

    def get(self, request, *args, **kwargs):
        status = self.request.GET.get("status", "")
        role = self.request.GET.get("role", "")
        query = self.request.GET.get("q", "")
        users = CustomUser.objects.order_by("full_name")
        users = utils.get_filtered_users(users, status=status, role=role, query=query)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="User List.csv"'

        writer = csv.writer(response)
        writer.writerow(
            ["Name", "Status", "Role", "Email", "Phone", "Gender", "Date Joined"]
        )

        for user in users:
            writer.writerow(
                [
                    user.full_name,
                    "Active" if user.is_active else "Disabled",
                    "Admin" if user.is_superuser else "User",
                    user.email,
                    user.phone_number,
                    user.gender,
                    localtime(user.date_joined).strftime("%d-%m-%Y %I:%M:%S %p %Z"),
                ]
            )

        return response


class UserListPrintView(LoginRequiredMixin, MyPermissionRequiredMixin, TemplateView):
    """Print user list using client printers"""

    template_name = "user/list_print.html"
    permission_required = "user.can_view_others_profile"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        status = self.request.GET.get("status", "")
        role = self.request.GET.get("role", "")
        query = self.request.GET.get("q", "")
        users = CustomUser.objects.order_by("full_name")
        context["users"] = utils.get_filtered_users(
            users, status=status, role=role, query=query
        )
        return context


class UserListCopyView(LoginRequiredMixin, MyPermissionRequiredMixin, View):
    """Copy user list to clipboard"""

    permission_required = "user.can_view_others_profile"

    def get(self, request, *args, **kwargs):
        status = self.request.GET.get("status", "")
        role = self.request.GET.get("role", "")
        query = self.request.GET.get("q", "")
        users = CustomUser.objects.order_by("full_name")
        users = utils.get_filtered_users(users, status=status, role=role, query=query)

        # Tab-separated values for copy-paste
        lines = ["Name\tStatus\tRole\tEmail\tPhone\tGender\tDate Joined"]

        for user in users:
            line = "\t".join(
                [
                    user.full_name or "",
                    "Active" if user.is_active else "Disabled",
                    "Admin" if user.is_superuser else "User",
                    user.email or "",
                    user.phone_number or "",
                    user.gender or "",
                    localtime(user.date_joined).strftime("%d-%m-%Y %I:%M:%S %p %Z")
                    or "",
                ]
            )
            lines.append(line)

        user_data = "\n".join(lines)

        return render(request, "user/list_copy.html", {"user_data": user_data})


class UserView(
    LoginRequiredMixin,
    MyPermissionRequiredMixin,
    DetailView,
):
    """View user details"""

    template_name = "user/view.html"
    model = CustomUser
    context_object_name = "userobj"
    permission_required = "user.can_view_profile"
    another_user_permission_required = "user.can_view_others_profile"
    permission_denied_message = "You do not have permission to view this page"

    def get_object(self, queryset=None):
        # If pk is not provided, show the logged-in user's profile
        pk = self.kwargs.get("pk")
        if pk:
            return get_object_or_404(CustomUser, pk=pk)
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        target_user = self.get_object()
        permissions = target_user.user_permissions.select_related(
            "content_type"
        ).order_by("content_type__app_label", "content_type__model", "codename")

        grouped_permissions = {}
        for perm in permissions:
            key = f"{perm.content_type.app_label} | {perm.content_type.name}"
            grouped_permissions.setdefault(key, []).append(perm)

        context["grouped_permissions"] = grouped_permissions

        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer

        return context


class UserEditView(
    LoginRequiredMixin,
    MyPermissionRequiredMixin,
    UpdateView,
):
    """Edit user details"""

    model = CustomUser
    form_class = UserEditForm
    template_name = "user/edit.html"
    permission_required = "user.can_edit_profile"
    another_user_permission_required = "user.can_edit_others_profile"
    permission_denied_message = "You do not have permission to edit this user"

    def get_object(self, queryset=None):
        # If pk is not provided, show the logged-in user's profile
        pk = self.kwargs.get("pk")
        if pk:
            return get_object_or_404(CustomUser, pk=pk)

        return self.request.user

    def get_success_url(self):
        # Fetch the object to ensure the pk is available
        obj = self.get_object()
        return reverse_lazy("user_view", kwargs={"pk": obj.pk})

    def get_form(self, form_class=None):
        """
        Dynamically modify the form fields based on user permissions.
        """
        form = super().get_form(form_class)
        current_user = self.request.user
        target_user = self.get_object()

        # Hide sensitive fields if:
        # - the user is not a superuser, OR
        # - the user is a superuser editing their own profile

        if not current_user.is_superuser or current_user == target_user:
            for field in [
                "is_superuser",
                "is_staff",
                "is_active",
                "groups",
                "user_permissions",
            ]:
                if field in form.fields:
                    form.fields.pop(field)

        return form

    def form_valid(self, form):
        messages.success(
            self.request,
            f"User updated successfully",
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        permissions = Permission.objects.select_related("content_type").order_by(
            "content_type__app_label", "content_type__model", "codename"
        )

        grouped_permissions = {}
        for perm in permissions:
            key = f"{perm.content_type.app_label} | {perm.content_type.name}"
            grouped_permissions.setdefault(key, []).append(perm)

        context["grouped_permissions"] = grouped_permissions

        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer

        return context


class UserChangePasswordView(
    LoginRequiredMixin,
    MyPermissionRequiredMixin,
    PasswordChangeView,
):
    """Change user passowrd"""

    form_class = CustomPasswordChangeForm
    template_name = "user/change_password.html"
    permission_required = "user.can_change_password"
    another_user_permission_required = "user.can_change_others_password"
    permission_denied_message = "You do not have permission to change password"

    def get_object(self, queryset=None):
        # If pk is not provided, show the logged-in user's profile
        pk = self.kwargs.get("pk")
        if pk:
            return get_object_or_404(CustomUser, pk=pk)

        return self.request.user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.get_object()
        kwargs["target_user"] = self.request.user
        return kwargs

    def get_success_url(self):
        # Fetch the object to ensure the pk is available
        obj = self.get_object()
        return reverse_lazy("user_edit", kwargs={"pk": obj.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        userobj = self.get_object()
        context["userobj"] = userobj
        return context


class UserAddView(LoginRequiredMixin, MyPermissionRequiredMixin, CreateView):
    """Adds new user"""

    model = CustomUser
    form_class = UserAddForm
    template_name = "user/add.html"
    success_url = reverse_lazy("user_list")
    permission_required = "user.can_add_user"
    permission_denied_message = "You do not have permission to add user"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer
        return context

    def form_valid(self, form):
        form.instance.password = make_password(form.cleaned_data["password1"])
        messages.success(
            self.request,
            f'User "{form.cleaned_data["full_name"]}" added successfully',
        )
        return super().form_valid(form)


class UserDeleteView(LoginRequiredMixin, MyPermissionRequiredMixin, DeleteView):
    """Deletes the user"""

    model = CustomUser
    template_name = "common/confirm_delete.html"  # Template for confirmation
    success_url = reverse_lazy("user_list")
    permission_required = "user.can_delete_user"
    permission_denied_message = "You do not have permission to delete this user"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["delete_message"] = (
            f'Are you sure you want to delete the "{self.get_object().full_name}" ?'
        )
        context["cancel_url"] = "user_list"
        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'User "{self.object.full_name}" has been deleted.')
        return super().post(request, *args, **kwargs)

    def test_func(self):
        """
        Allow delete if the user is:
        - Not self
        - A superuser or,
        - Has permission_required
        """

        # Prevent self deletion
        if self.request.user == self.get_object():
            self.permission_denied_message = (
                "You do not have permission to delete yourself"
            )
            return False

        return super().test_func()


class UserGroupView(
    LoginRequiredMixin,
    MyPermissionRequiredMixin,
    DetailView,
):
    """View user group details"""

    template_name = "user/group_view.html"
    model = Group
    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to view the user groups"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.object

        # Organize permissions by content type for better display
        permissions = group.permissions.select_related("content_type").order_by(
            "content_type__app_label", "content_type__model", "codename"
        )

        grouped_permissions = {}
        for perm in permissions:
            key = f"{perm.content_type.app_label} | {perm.content_type.name}"
            grouped_permissions.setdefault(key, []).append(perm)

        context["grouped_permissions"] = grouped_permissions

        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer

        return context


class UserGroupListView(MyPermissionRequiredMixin, ListView):
    "View user group list"

    model = Group
    template_name = "user/group_list.html"
    context_object_name = "groups"
    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to view the user groups"
    paginate_by = 10

    def get_paginate_by(self, queryset):
        page_size = self.request.GET.get("page_size")
        if page_size and page_size.isdigit():
            return int(page_size)
        return self.paginate_by

    def get_queryset(self):

        query = self.request.GET.get("q", "")
        queryset = Group.objects.all().order_by("name")

        if query:
            queryset = queryset.filter(name__icontains=query)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["current_page_size"] = self.get_paginate_by(self.get_queryset())
        context["page_size_options"] = [5, 10, 20, 50, 100]
        return context


class UserGroupListExportAsXlsxView(
    LoginRequiredMixin, MyPermissionRequiredMixin, View
):
    """Export user group list as .xlsx file"""

    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to perform this action"

    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")
        queryset = Group.objects.all().order_by("name")

        if query:
            queryset = queryset.filter(name__icontains=query)

        wb = Workbook()
        ws = wb.active or wb.create_sheet(title="Sheet1")
        ws.append(["Group", "Permissions", "Users"])

        for group in queryset:
            ws.append(
                [
                    group.name,
                    group.permissions.count(),
                    group.user_set.count(),
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Group List.xlsx"'

        wb.save(response)
        return response


class UserGroupListExportAsCsvView(LoginRequiredMixin, MyPermissionRequiredMixin, View):
    """Export user group list as .csv file"""

    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to perform this action"

    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")
        queryset = Group.objects.all().order_by("name")

        if query:
            queryset = queryset.filter(name__icontains=query)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="Group List.csv"'

        writer = csv.writer(response)
        writer.writerow(["Group", "Permissions", "Users"])

        for group in queryset:
            writer.writerow(
                [
                    group.name,
                    group.permissions.count(),
                    group.user_set.count(),
                ]
            )

        return response


class UserGroupListPrintView(
    LoginRequiredMixin, MyPermissionRequiredMixin, TemplateView
):
    """Print user group list using client side printers"""

    template_name = "user/group_list_print.html"
    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to perform this action"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        query = self.request.GET.get("q", "")
        queryset = Group.objects.all().order_by("name")

        if query:
            queryset = queryset.filter(name__icontains=query)

        context["groups"] = queryset
        return context


class UserGroupListCopyView(LoginRequiredMixin, MyPermissionRequiredMixin, View):
    """Copy user group list to clipboard"""

    permission_required = "auth.view_group"
    permission_denied_message = "You do not have permission to perform this action"

    def get(self, request, *args, **kwargs):

        query = self.request.GET.get("q", "")
        queryset = Group.objects.all().order_by("name")

        if query:
            queryset = queryset.filter(name__icontains=query)

        # Tab-separated values for copy-paste
        lines = ["Group\tPermissions\tUsers"]

        for group in queryset:
            line = "\t".join(
                [
                    group.name or "",
                    str(group.permissions.count()) or "",
                    str(group.user_set.count()) or "",
                ]
            )
            lines.append(line)

        group_data = "\n".join(lines)

        return render(request, "user/group_list_copy.html", {"group_data": group_data})


class UserGroupAddView(LoginRequiredMixin, MyPermissionRequiredMixin, CreateView):
    """Add a group along with permissions in one step"""

    model = Group
    form_class = UserGroupAddForm
    template_name = "user/group_add_edit.html"
    permission_required = "auth.add_group"
    permission_denied_message = "You do not have permission to add a group"
    success_url = reverse_lazy("user_group_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Add New Group"
        context["button_text"] = "Add Group"

        permissions = Permission.objects.select_related("content_type").order_by(
            "content_type__app_label", "content_type__model", "codename"
        )

        grouped_permissions = {}
        for perm in permissions:
            key = f"{perm.content_type.app_label} | {perm.content_type.name}"
            grouped_permissions.setdefault(key, []).append(perm)

        context["grouped_permissions"] = grouped_permissions

        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer

        return context

    def form_valid(self, form):
        # Create the group
        self.object = Group.objects.create(name=form.cleaned_data["name"])

        # Assign permissions
        permissions = form.cleaned_data["permissions"]
        if permissions:
            self.object.permissions.set(permissions)

        perm_count = permissions.count() if permissions else 0
        perm_word = "permission" if perm_count == 1 else "permissions"

        messages.success(
            self.request,
            f'Group "{self.object.name}" created with {perm_count} {perm_word}.',
        )
        return redirect(self.success_url)


class UserGroupEditView(LoginRequiredMixin, MyPermissionRequiredMixin, UpdateView):
    """Edit a group along with permissions in one step"""

    model = Group
    form_class = UserGroupAddForm
    template_name = "user/group_add_edit.html"
    permission_required = "auth.change_group"
    permission_denied_message = "You do not have permission to edit this group"

    def get_success_url(self):
        # Fetch the object to ensure the pk is available
        obj = self.get_object()
        return reverse_lazy("user_group_view", kwargs={"pk": obj.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Edit Group"
        context["button_text"] = "Save Changes"

        permissions = Permission.objects.select_related("content_type").order_by(
            "content_type__app_label", "content_type__model", "codename"
        )

        grouped_permissions = {}
        for perm in permissions:
            key = f"{perm.content_type.app_label} | {perm.content_type.name}"
            grouped_permissions.setdefault(key, []).append(perm)

        context["grouped_permissions"] = grouped_permissions

        referer = self.request.META.get("HTTP_REFERER")
        context["referer"] = referer

        return context

    def form_valid(self, form):
        self.object = form.save()

        # Assign permissions
        permissions = form.cleaned_data["permissions"]
        if permissions:
            self.object.permissions.set(permissions)
        else:
            self.object.permissions.clear()

        perm_count = permissions.count() if permissions else 0
        perm_word = "permission" if perm_count == 1 else "permissions"

        messages.success(
            self.request,
            f'Group "{self.object.name}" updated with {perm_count} {perm_word}.',
        )
        return redirect(self.get_success_url())


class UserGroupDeleteView(LoginRequiredMixin, MyPermissionRequiredMixin, DeleteView):
    """Delete the group"""

    model = Group
    template_name = "common/confirm_delete.html"  # Template for confirmation
    success_url = reverse_lazy("user_group_list")
    permission_required = "auth.delete_group"
    permission_denied_message = "You do not have permission to delete this group"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["delete_message"] = (
            f'Are you sure you want to delete the "{self.get_object().name}" Group?'
        )
        context["cancel_url"] = "user_group_list"
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Check if any user is assigned to this group
        if self.object.user_set.exists():
            messages.error(
                request,
                f'Cannot delete group "{self.object.name}" because it has users assigned.',
            )
            return redirect(self.success_url)

        # Safe to delete
        messages.success(request, f'Group "{self.object.name}" has been deleted.')
        return super().post(request, *args, **kwargs)
